from pathlib import Path
from typing import Dict, List, Tuple, Optional
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from matching.exceptions import ExcelProcessingError


class ExcelProcessor:
    """
    Klasa odpowiedzialna za wszystkie operacje na plikach Excel.
    Implementuje wzorzec Singleton, aby zapewnić jeden punkt dostępu do otwartych plików.
    """

    def __init__(self):
        # Słownik przechowujący otwarte skoroszyty {ścieżka: workbook}
        self.workbooks: Dict[str, openpyxl.Workbook] = {}

        # Maksymalne limity dla bezpieczeństwa
        self.MAX_FILE_SIZE_MB = 10
        self.MAX_SHEETS = 10

    def load_files(self, working_file: Path, reference_file: Path) -> None:
        """
        Wczytuje pliki Excel do pamięci.

        Args:
            working_file: Ścieżka do pliku roboczego (WF)
            reference_file: Ścieżka do pliku referencyjnego (REF)

        Raises:
            ExcelProcessingError: Gdy wystąpi problem z wczytaniem plików
        """
        print("DEBUG: *** load_files *** was called from the ExcelProcessor")

        try:
            # Zamknij poprzednio otwarte pliki
            self.close_all_workbooks()

            # Wczytaj nowe pliki
            for file_path in [working_file, reference_file]:
                if not file_path.exists():
                    raise ExcelProcessingError(f"Plik nie istnieje: {file_path}")

                # Sprawdź rozmiar pliku
                file_size_mb = file_path.stat().st_size / (1024 * 1024)
                if file_size_mb > self.MAX_FILE_SIZE_MB:
                    raise ExcelProcessingError(
                        f"Plik {file_path} przekracza maksymalny rozmiar {self.MAX_FILE_SIZE_MB}MB"
                    )

                # Wczytaj plik
                workbook = openpyxl.load_workbook(file_path, data_only=True)

                # Sprawdź liczbę arkuszy
                if len(workbook.sheetnames) > self.MAX_SHEETS:
                    raise ExcelProcessingError(
                        f"Plik {file_path} ma zbyt wiele arkuszy (max: {self.MAX_SHEETS})"
                    )

                self.workbooks[str(file_path)] = workbook

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas wczytywania plików: {str(e)}")

    def read_descriptions(
        self, file_path: Path, column: str, cell_range: Dict[str, str]
    ) -> List[Tuple[str, str]]:
        """
        Czyta opisy z określonej kolumny i zakresu.

        Args:
            file_path: Ścieżka do pliku Excel
            column: Litera kolumny (np. 'A', 'B')
            cell_range: Słownik z kluczami 'start' i 'end' określającymi zakres

        Returns:
            Lista krotek (opis, adres_komórki)

        Raises:
            ExcelProcessingError: Gdy wystąpi problem z odczytem danych
        """
        print("DEBUG: *** read_descriptions *** was called from the ExcelProcessor")

        try:
            workbook = self.workbooks[str(file_path)]
            sheet = workbook.active

            # Pobierz numery wierszy z zakresu
            start_row = int(cell_range["start"])
            end_row = int(cell_range["end"])

            descriptions = []
            for row in range(start_row, end_row + 1):
                cell_address = f"{column}{row}"
                cell_value = sheet[cell_address].value

                # Pomiń puste komórki
                if cell_value is not None:
                    descriptions.append((str(cell_value).strip(), cell_address))

            return descriptions

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas odczytu opisów: {str(e)}")

    def read_prices(
        self, file_path: Path, price_column: str, row_range: Dict[str, str]
    ) -> Dict[str, Decimal]:
        """
        Czyta ceny z określonej kolumny i zakresu.

        Args:
            file_path: Ścieżka do pliku Excel
            price_column: Litera kolumny z cenami
            row_range: Słownik z kluczami 'start' i 'end' określającymi zakres

        Returns:
            Słownik {adres_komórki: cena}

        Raises:
            ExcelProcessingError: Gdy wystąpi problem z odczytem lub konwersją cen
        """
        print("DEBUG: *** read_prices *** was called from the ExcelProcessor")

        try:
            workbook = self.workbooks[str(file_path)]
            sheet = workbook.active

            # Pobierz numery wierszy z zakresu
            start_row = int(row_range["start"])
            end_row = int(row_range["end"])

            prices = {}
            for row in range(start_row, end_row + 1):
                cell_address = f"{price_column}{row}"
                cell_value = sheet[cell_address].value

                # Pomiń puste komórki
                if cell_value is not None:
                    try:
                        # Konwersja na Decimal dla precyzji finansowej
                        price = Decimal(str(cell_value))
                        prices[cell_address] = price
                    except (ValueError, TypeError, Decimal.InvalidOperation):
                        raise ExcelProcessingError(
                            f"Nieprawidłowa wartość ceny w komórce {cell_address}"
                        )

            return prices

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas odczytu cen: {str(e)}")

    def write_price(self, file_path: str, cell_address: str, price: Decimal) -> None:
        """
        Zapisuje cenę do określonej komórki.

        Args:
            file_path: Ścieżka do pliku Excel
            cell_address: Adres komórki (np. 'F5')
            price: Cena do zapisania

        Raises:
            ExcelProcessingError: Gdy wystąpi problem z zapisem
        """
        print("DEBUG: *** write_price *** was called from the ExcelProcessor")

        try:
            workbook = self.workbooks[file_path]
            sheet = workbook.active
            sheet[cell_address] = float(price)  # Konwersja na float dla Excel

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas zapisu ceny: {str(e)}")

    def close_all_workbooks(self) -> None:
        """
        Zamyka wszystkie otwarte pliki Excel.
        """
        print("*** close_all_workbooks *** was called from the ExcelProcessor")

        for workbook in self.workbooks.values():
            workbook.close()
        self.workbooks.clear()

    def __del__(self):
        """
        Destruktor - upewnia się, że wszystkie pliki zostały zamknięte.
        """
        self.close_all_workbooks()
