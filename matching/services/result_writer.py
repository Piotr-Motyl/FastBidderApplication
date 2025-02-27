from typing import List, Dict
from decimal import Decimal
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from matching.exceptions import ExcelProcessingError


class ResultWriter:
    """Serwis odpowiedzialny za zapisanie wyników porównania i cen"""

    # Stałe do formatowania
    SOURCE_INFO_COLUMN_HEADER = "Źródło ceny"
    REPORT_FILENAME_TEMPLATE = "matching_report_{timestamp}.xlsx"
    HIGHLIGHT_FILL = PatternFill(
        start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"
    )

    def __init__(self, excel_processor):
        self.excel_processor = excel_processor

    def write_results(
        self, results: List[Dict], working_file_path: Path, price_target_column: str
    ) -> str:
        """Zapisuje wyniki dopasowania do pliku WF i generuje raport

        Args:
            results (List[Dict]): Lista słowników z wynikami dopasowania
            working_file_path (Path): Ścieżka do pliku WF

        Returns:
            str: Ścieżka do wygenerowanego pliku raportu

        Raises:
            ExcelProcessingError: W przypadku błędów podczas zapisu do pliku
        """
        print("*** wywołano metodę ===write_results=== z ResultWriter")

        try:
            # DEBUG - sprawdź zawartość results
            print(f"Liczba znalezionych dopasowań: {len(results)}")
            for i, result in enumerate(results[:3]):
                print(f"Dopasowanie {i+1}: {result}")

            # Upewnij się, że plik istnieje
            if not working_file_path.exists():
                raise ExcelProcessingError(f"Plik {working_file_path} nie istnieje")

            # Zapisz wyniki do pliku roboczego
            self._write_to_working_file(results, working_file_path, price_target_column)

            # Wygeneruj raport
            report_path = self._generate_report(
                results, working_file_path, price_target_column
            )

            return str(report_path)

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas zapisywania wyników: {str(e)}")

    def _write_to_working_file(
        self, results: List[Dict], file_path: Path, price_target_column: str
    ) -> None:
        """Zapisuje ceny i informacje o źródle do pliku WF

        Args:
            results (List[Dict]): Lista wyników dopasowania
            file_path (Path): Ścieżka do pliku roboczego
            price_target_column (str): Kolumna docelowa dla cen
        """
        print("*** wywołano metodę ===_write_to_working_file=== z ResultWriter")
        file_path_str = str(file_path)  # Konwersja Path na string dla ExcelProcessor

        try:
            # Otwórz plik jeśli nie jest jeszcze otwarty przez ExcelProcessor
            if file_path_str not in self.excel_processor.workbooks:
                workbook = openpyxl.load_workbook(file_path_str)
                self.excel_processor.workbooks[file_path_str] = workbook
            else:
                workbook = self.excel_processor.workbooks[file_path_str]

            sheet = workbook.active

            # Znajdź lub utwórz kolumnę na informacje o źródle
            source_info_col = self._get_or_create_source_info_column(sheet)

            # Zapisz wyniki
            for result in results:
                # Pobierz dane z słownika wynikowego
                wf_cell = result["wf_cell"]
                price = result["price"]

                # Określ komórkę docelową dla ceny używając price_target_column
                cell_row = wf_cell[1:]  # Pobierz numer wiersza z adresu komórki
                price_target_cell = f"{price_target_column}{cell_row}"

                # Zapis ceny bezpośrednio do arkusza
                sheet[price_target_cell] = float(price) if price else 0.0

                # Zapis informacji o źródle w kolumnie informacyjnej
                source_cell = f"{source_info_col}{cell_row}"
                source_info = f"REF:{result['ref_cell']}, Podobieństwo: {result['match_score']:.1f}%"

                sheet[source_cell] = source_info
                sheet[source_cell].fill = self.HIGHLIGHT_FILL

            # Zapisz zmiany do pliku
            workbook.save(file_path_str)

        except Exception as e:
            raise ExcelProcessingError(f"Błąd podczas zapisu do pliku: {str(e)}")

    def _generate_report(
        self, results: List[Dict], working_file_path: Path, price_target_column: str
    ) -> str:
        """Generuje szczegółowy raport dopasowań

        Args:
            results (List[Dict]): Lista wyników dopasowania
            working_file_path (Path): Ścieżka do pliku roboczego
            price_target_column (str): Kolumna docelowa dla cen

        Returns:
            str: Ścieżka do wygenerowanego raportu
        """
        print("*** wywołano metodę ===_generate_report=== z ResultWriter")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = working_file_path.parent / self.REPORT_FILENAME_TEMPLATE.format(
            timestamp=timestamp
        )

        # Tworzymy nowy plik Excel na raport
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Raport dopasowań"

        # Nagłówki
        headers = [
            "Opis WF",
            "Komórka WF",
            "Opis REF",
            "Komórka REF",
            "Cena",
            "Podobieństwo (%)",
            "Komórka docelowa ceny",
        ]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Wypełnienie danymi
        for row, result in enumerate(results, 2):
            # Określ komórkę docelową dla ceny używając price_target_column
            cell_row = result["wf_cell"][1:]  # Pobierz numer wiersza z adresu komórki
            price_target_cell = f"{price_target_column}{cell_row}"

            sheet.cell(row=row, column=1, value=result["wf_description"])
            sheet.cell(row=row, column=2, value=result["wf_cell"])
            sheet.cell(row=row, column=3, value=result["ref_description"])
            sheet.cell(row=row, column=4, value=result["ref_cell"])
            sheet.cell(row=row, column=5, value=float(result["price"]))
            sheet.cell(row=row, column=6, value=round(result["match_score"], 1))
            sheet.cell(row=row, column=7, value=price_target_cell)

        workbook.save(report_path)
        return str(report_path)

    def _get_or_create_source_info_column(self, sheet) -> str:
        """
        Znajduje istniejącą lub tworzy nową kolumnę na informacje o źródle

        Args:
            sheet: Arkusz Excel

        Returns:
            str: Litera kolumny (np. 'G')
        """
        print(
            "*** wywołano metodę ===_get_or_create_source_info_column=== z ResultWriter"
        )
        # Szukaj istniejącej kolumny
        for cell in sheet[1]:
            if cell.value == self.SOURCE_INFO_COLUMN_HEADER:
                return cell.column_letter

        # Utwórz nową kolumnę
        last_col = sheet.max_column
        new_col_letter = openpyxl.utils.get_column_letter(last_col + 1)
        sheet[f"{new_col_letter}1"] = self.SOURCE_INFO_COLUMN_HEADER
        return new_col_letter
